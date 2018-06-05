import pandas as pd
import Simple_control as sp
import time
import Complex_control as cp
from datetime import datetime
import openpyxl as op
from openpyxl.utils.dataframe import dataframe_to_rows


class Control_center:
    def __init__(self, path_datasource, worksheets):
        self.path_datasource = path_datasource  # containing the path of the datasource
        self.datasource = pd.DataFrame()    # DataFrame containing the analysed dataset
        self.bool_result = pd.DataFrame()   # DataFrame containing the boolean result
        self.text_result = pd.DataFrame()    # DataFrame containing the commented result
        self.worksheets = worksheets    # worksheets of the Excel source file
        self.__load_excel_datasource()  # transfer of the excel file containing data to source DataFrame
        self.nb_fistrows = 0

    def __load_excel_datasource(self):

        try:
            self.Excel = pd.ExcelFile(self.path_datasource)
            self.datasource = self.Excel.parse(self.worksheets)
        except:
            print("the excel file doest exist")

    def first_raws(self, *args):
        self.nb_fistrows = 0
        try:
            for arg in args:
                self.bool_result[arg] = self.datasource[arg]
                self.text_result[arg] = self.datasource[arg]
                self.nb_fistrows += 1
        except:
            print("the column doest exist")

        if self.nb_fistrows != 0:
            cols = self.bool_result.columns.tolist()
            total_cols = len(cols) - self.nb_fistrows
            cols = cols[-self.nb_fistrows:] + cols[:total_cols]
            self.bool_result = self.bool_result[cols]

            cols = self.text_result.columns.tolist()
            total_cols = len(cols) - self.nb_fistrows
            cols = cols[-self.nb_fistrows:] + cols[:total_cols]
            self.text_result = self.text_result[cols]

    def convert_col(self,column_name, format, newname = None):
        if newname == None:
            if format == "Date":
                self.datasource[column_name] = pd.to_datetime(self.datasource[column_name], errors='coerce')

            if format == "String":
                #self.datasource[column_name] = pd.DataFrame.to_string(self.datasource, columns=[column_name])
                self.datasource[column_name] = self.datasource[column_name].astype('str')
        else:
            if format == "Date":
                self.datasource[newname] = pd.to_datetime(self.datasource[column_name], errors='coerce')

            if format == "String":
                # self.datasource[newname] = pd.DataFrame.to_string(self.datasource, columns=[column_name])
                self.datasource[newname] = self.datasource[column_name].astype('str')

    def empty(self, column_name, showed, control_name="Empty_Test", error_message="Empty", reverse="F"):
        control_empty = sp.Simple_control(control_name,  column_name, error_message, self.datasource, showed,reverse)
        control_empty.boolean_control = pd.isna(self.datasource[column_name]) # Check empty cells and update bool_result list
        self.update_DataFrame(control_empty)

    def datecompare(self,date1,date2,compare,showed,control_name="Date_compare", error_message="Empty", reverse="F"):
        control_datecomp = sp.Simple_control(control_name, date1,error_message,self.datasource,showed,reverse)

        if "/" in date2:
            date2 = datetime.strptime(date2 ,'%d/%m/%Y')
            for i, val in enumerate(self.datasource.itertuples()):
                if compare == ">":
                    self.datasource.at[i, "delta_" + control_name] = self.datasource.at[i, date1] - date2
                    if self.datasource.at[i, "delta_" + control_name].days > 0:
                        control_datecomp.boolean_control.append(True)
                    else:
                        control_datecomp.boolean_control.append(False)
                if compare == "<":
                    self.datasource.at[i, "delta_" + control_name] = self.datasource.at[i, date2] - date2
                    if self.datasource.at[i, "delta_" + control_name].days > 0:
                        control_datecomp.boolean_control.append(True)
                    else:
                        control_datecomp.boolean_control.append(False)
        else:
            for i,val in enumerate(self.datasource.itertuples()):
                if compare == ">":
                    self.datasource.at[i, "delta_"+control_name] = self.datasource.at[i, date1]-self.datasource.at[i, date2]
                    if self.datasource.at[i, "delta_"+control_name].days > 0:
                        control_datecomp.boolean_control.append(True)
                    else:
                        control_datecomp.boolean_control.append(False)

                if compare == "<":
                    self.datasource.at[i, "delta_"+control_name] = self.datasource.at[i, date2]-self.datasource.at[i, date1]
                    if self.datasource.at[i, "delta_"+control_name].days > 0:
                        control_datecomp.boolean_control.append(True)
                    else:
                        control_datecomp.boolean_control.append(False)

        self.update_DataFrame(control_datecomp)

    def Isequal(self, column_name,showed,control_name="Equality_Test", error_message="wrong value",reverse="F",*values):
        control_isequal= sp.Simple_control(control_name,  column_name, error_message, self.datasource, showed,reverse)
        control_isequal.boolean_control = self.datasource[column_name].apply(lambda x : True if x in values else False)
        self.update_DataFrame(control_isequal)

    def begwith(self,column_name,showed,control_name="Beg_Test", error_message="wrong value",reverse="F",*values):
        control_begwith = sp.Simple_control(control_name,  column_name, error_message, self.datasource, showed,reverse)
        control_begwith.boolean_control = self.datasource[column_name].apply(lambda x: True if (x.startswith(values)) is True else False)
        self.update_DataFrame(control_begwith)

    def space(self,column_name,showed,position, control_name="Space_Anomaly",error_message="Space",reverse="F"):

        control_space = sp.Simple_control(control_name, column_name, error_message, self.datasource, showed,reverse)

        if position =="beg":

            control_space.boolean_control = self.datasource[column_name].apply(lambda x: True if x[0] == " " else False)


        if position =="end":
            control_space.boolean_control = self.datasource[column_name].apply(lambda x: print if x[-1] == " " else False)

        self.update_DataFrame(control_space)

    def update_DataFrame(self, control):

        if not isinstance(control, sp.Simple_control) and not isinstance(control, cp.Complex_control): #Load bool result in the DataFrame
            raise TypeError("control must be a Simple_control or a Complex_control")

        if control.control_name not in self.bool_result.columns:
            self.bool_result.insert(0, str(control.control_name), "")

        for i, cel in enumerate(control.boolean_control):
            if control.reverse == "F":
                self.bool_result.at[i, control.control_name] = cel
            if control.reverse == "T":
                if cel is False:
                    self.bool_result.at[i, control.control_name] = True
                if cel is True:
                    self.bool_result.at[i, control.control_name] = False

        if control.showed == 1:     #Load text result in the DataFrame
            if control.control_name not in self.text_result.columns:
                self.text_result.insert(0, str(control.control_name), "")
                control.build_text_result()

            for i, cel in enumerate(control.commented_anomaly):
                self.text_result.at[i, control.control_name] = cel

    def list_columns(self):
        print(self.bool_result.columns)

    def complex_control(self,control_name, error_message, control_validation,showed):
        comp_c = cp.Complex_control(control_name, error_message, self.bool_result,control_validation,showed)
        self.update_DataFrame(comp_c)

    def __synthesis(self):
        columns = self.text_result.columns
        columns = columns[self.nb_fistrows:]
        self.text_result["Synthèse"] = ""
        for col in columns:
            for i,val in enumerate(self.text_result[col]):
                if val != "":
                    self.text_result.at[i, "Synthèse"] = self.text_result.at[i, "Synthèse"] + str(val) + "/"

        cols = self.text_result.columns.tolist()
        total_cols = len(cols) - self.nb_fistrows
        cols = cols[-1:] + cols[:total_cols]
        self.text_result = self.text_result[cols]

    def export_excel(self,directory,result_type,sheet="Results"):
        """Méthode permettant d'extraire le résultat sous forme d'un fichier excel"""
        self.__synthesis()
        if result_type == "bool":
            writer = pd.ExcelWriter(directory, engine='xlsxwriter')
            self.bool_result.to_excel(writer, sheet_name=sheet)
            writer.save()
        if result_type == "text":
            writer = pd.ExcelWriter(directory, engine='xlsxwriter')

            self.text_result.to_excel(writer, sheet_name=sheet)

    def toexcel(self, directory):

        """toexcel export dataframe into an Excel File."""
        #Detailed view
        self.__synthesis()
        wb = op.Workbook()
        ws = wb.active
        for r in dataframe_to_rows(self.text_result, index=True, header=True):
            ws.append(r)
        ws.delete_cols(1, amount=1)
        ws.delete_rows(2, amount=1)
        ws.title = "Liste détaillée"
        ws.column_dimensions["A"].width = 45

        cols = self.text_result.columns.tolist()
        for i, val in enumerate(cols):
            if i not in [0,1]:
                if len(val) < 15:
                    ws.column_dimensions[op.utils.get_column_letter(i+1)].width = 15
                else:
                    ws.column_dimensions[op.utils.get_column_letter(i+1)].width = len(val)
                    #ws[op.utils.get_column_letter(i+1)+1].p
                    print(op.utils.get_column_letter(i),len(val), val)
        wb.save(directory)


t3 = time.clock()
montest = Control_center(r'D:\Users\sgasmi\Desktop\Données maquette\Data_model - Copie.xlsx', 'Feuil1')
t4 = time.clock()
#montest = Control_center(r'C:\Users\Sabri.GASMI\Desktop\Jeu - Copie.xlsx', 'Feuil1')

t5 = time.clock()
montest.convert_col("Prénom","String")
montest.convert_col("Matricule","String","str_Matricule")
montest.convert_col("Date d'entrée société","Date")
montest.convert_col("Date d'entrée groupe","Date")
t6 = time.clock()

#print(montest.datasource.dtypes)
t1 = time.clock()
#conditions
montest.begwith("str_Matricule",0,"Scope_FR","FR","F","6")
montest.Isequal("Tranche décompte",0,"Scope_decompte","Scope_Decompte","T",1,99)
montest.Isequal("Clé statut d'activité",0,"Scope_activite","Scope_activite","F",1,3)
montest.Isequal("CtSAL",0,"Scope_ct_SAL","Scope_ct_SAL","F",1,8)
#test simples
montest.datecompare("Date d'entrée société","Date d'entrée groupe",">",1,"Ecart date","Date entrée invalide")
montest.datecompare("Date de naissance","30/06/1981",">",1,"Ecart age", "Age improbable")
montest.empty("Matricule",1,"Matricule_vide","Matricule vide")
montest.empty("Prénom",0,"Prénom_vide","Prénom vide")
montest.empty("Nom",0,"Nom_vide","Nom vide")
montest.empty("Sexe",0,"Sexe_vide","Sexe vide")
montest.empty("Clé situation de famille",0,"Clé_famille_vide","Clé situation de famille vide")
montest.empty("Date de naissance",0,"Date_naissance_vide","Date de naissance vide")
montest.empty("Nationalité",0,"Nationalité_vide","Nationalité vide")
montest.empty("Type de contrat",0,"Type_contrat_vide","Type de contrat vide")
montest.empty("Type conv. collec.",0,"Type_conv_coll_vide","Type de convention coll vide")
montest.empty("Statut de salariés",0,"Statut_de_salariés_vide")
montest.empty("Statut",0,"Statut_vide")
montest.empty("DPer",0,"Dper_vide")
montest.empty("Libellé société",0,"Libellé_société_vide")
montest.empty("Domaine du personnel",0,"Domaine_du_pers_vide")
montest.empty("Niveau",0,"Niveau_vide")
montest.empty("Classif",0,"Classif_vide")
montest.empty("Qualification bulletin",0,"Qualif_bulletin_vide")
montest.empty("Date d'entrée société", 0,"Date_entrée_société")
montest.empty("Délégation",0,"Délégation_vide")
montest.empty("Région",0,"Région_vide")
montest.empty("Agence",0,"Agence_vide")
montest.space("Nom",1,"end","Nom_beg")

#tests complexe
montest.complex_control("Conditions","Ok","Scope_FR*Scope_decompte*Scope_activite*Scope_activite",1)
montest.complex_control("Matricule vide","Matricule vide","Matricule_vide*Conditions",1)
montest.complex_control("Prénom vide","Conditions","Prénom_vide*Conditions", 1)
montest.complex_control("Nom vide","Nom vide","Nom_vide*Conditions",1)
montest.complex_control("Sexe vide", "Sexe vide", "Sexe_vide*Conditions",1)
montest.complex_control("Type de contrat vide","Type de contrat vide","Type_contrat_vide*Conditions",1)
montest.complex_control("Nationalité vide","Nationnalité vide", "Nationalité_vide*Conditions",1)
montest.complex_control("Conv collective vide","Type_conv_coll_vide","Type_conv_coll_vide*Conditions",1)
montest.complex_control("Statut de salariés","Statut de salarié vide","Statut_de_salariés_vide*Conditions",1)
montest.complex_control("Statut vide","Statut empty","Statut_vide*Conditions",1)
montest.complex_control("DPER vide","Dper vide","Dper_vide*Conditions",1)
montest.complex_control("Libellé société vide", "Libellé de société vide","Libellé_société_vide*Conditions",1)
montest.complex_control("Domaine du personnel vide", "Domaine du personnel vide", "Domaine_du_pers_vide*Conditions",1)
montest.complex_control("Niveau vide","Niveau vide", "Niveau_vide*Conditions",1)
montest.complex_control("Classif vide", "Classif vide", "Classif_vide*Conditions",1)
montest.complex_control("Délégation","Délégation vide","Délégation_vide*Conditions",1)
montest.complex_control("Région","Région vide","Région_vide*Conditions",1)
montest.complex_control("Agence","Agence vide","Agence_vide*Conditions",1)


montest.first_raws("Matricule", "Nom")

#montest.complex_control("Matricule vide")

t2 = time.clock()
t7 = time.clock()

montest.toexcel(r"D:\Users\sgasmi\Desktop\toto.xlsx")
#montest.export_excel(r'D:\Users\sgasmi\Desktop\Anomalies.xlsx',"text")

t8 = time.clock()

print(t2 - t1,"tests")
print(t4 - t3, "chargement")
print(t6-t5, "conversion")
print(t8-t7,"Export")


#test = cp.Complex_control("TestC","Erreur",montest.bool_result,"Nom*Prénom",0)
"""
montest.empty("Date de naissance", 1, "Date naiss", "Date naissance vide")
montest.empty("Clé situation de famille", 1, "Ctr Clé situation de famille", "Situation vide",)
montest.first_raws("Matricule", "Nom")
"""

"""
cols = df.columns.tolist()
>>> cols = [cols[-1]]+cols[:-1] # or whatever change you need
>>> df.reindex(columns=cols)"""